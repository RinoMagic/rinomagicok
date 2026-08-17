"""Excel Parser — canonical `.xlsx` ingestion for RinoMagic.

This module replaces the (slower, error-prone) PDF/OCR pipeline used by:
  * ``scoreandlive._parse_listone_pdf`` — the pre-season roster (Listone)
  * ``matchday_facts._parse_voti_pdf``   — the weekly voti (matchday facts)

Two entry points:

* :func:`parse_listone_xlsx`  → returns a list of player dicts ready to be
  inserted into the ``sal_players`` collection.
* :func:`parse_voti_xlsx`     → returns ``(matchday, rows, diagnostics)``
  where each row has the same shape produced by ``_parse_voti_pdf`` so the
  downstream storage / settlement code is untouched.

Input files come from https://www.fantacalcio.it/ (both are official Excel
exports the user already downloads today). Structure summary:

**Listone (``Quotazioni_Fantacalcio_Stagione_YYYY_YY``)** — sheet "Tutti"::

    R1: <title spanning A>
    R2: Id | R | RM | Nome | Squadra | Qt.A | Qt.I | Diff. | Qt.A M | Qt.I M | Diff.M | FVM | FVM M
    R3+: player rows (P, D, C, A only — ALL/allenatore absent here)

**Voti giornata (``Voti_Fantacalcio_Stagione_YYYY_YY_Giornata_N``)** —
sheet "Fantacalcio" (also "Statistico" and "Italia" variants)::

    R1: "Voti Fantacalcio Nª giornata di campionato"
    R2-R4: disclaimers
    R5: team header (single string in column A, e.g. "Atalanta")
    R6: Cod. | Ruolo | Nome | Voto | Gf | Gs | Rp | Rs | Rf | Au | Amm | Esp | Ass
    R7+: player rows, followed by the next team header, and so on

Autogol note (verified on real data): the goalkeeper's ``Gs`` column already
includes autogol conceded by teammates. No manipulation is required at fact
level — the sum ``Σ(Gf+Rf) + Σ(Au) == Σ(GK Gs)`` holds across the matchday.
"""
from __future__ import annotations

import io
import re
import logging
from typing import List, Tuple, Optional, Dict, Any

logger = logging.getLogger("excel_parser")

# --- Team canonicalisation (kept in sync with matchday_facts) -------------
# Import lazily to avoid a hard dependency on that module at import time in
# tests / scripts, but fall back to a self-contained copy if needed.
try:
    from matchday_facts import SERIE_A_TEAMS, TEAM_ALIASES  # type: ignore
except Exception:  # pragma: no cover
    SERIE_A_TEAMS = {
        "Atalanta", "Bologna", "Cagliari", "Como", "Cremonese", "Fiorentina",
        "Genoa", "Inter", "Juventus", "Lazio", "Lecce", "Milan", "Napoli",
        "Parma", "Pisa", "Roma", "Sassuolo", "Torino", "Udinese", "Verona",
        "Hellas Verona", "Empoli", "Monza", "Frosinone", "Salernitana",
        "Venezia", "Spezia", "Sampdoria",
    }
    TEAM_ALIASES = {"Hellas Verona": "Verona"}

_SERIE_A_LOWER = {t.lower(): t for t in SERIE_A_TEAMS}
_DISCLAIMER_MARKERS = (
    "Voti Fantacalcio",
    "Quotazioni Fantacalcio",
    "Calciatori Ceduti",
    "Voti Statistici",
    "Voti Italia",
    "fantacalcio.it",
    "QUESTO FILE",
    "ESCLUSIVO",
    "giornata",
    "Solo su www",
)
_MATCHDAY_RE = re.compile(r"(\d+)\s*[ªa°]?\s*giornata", re.IGNORECASE)
_INITIAL_RE = re.compile(r"^[A-Z]\.$")  # matches "L." / "N." / "Jo." handled below

_ROLE_ALLOWED_LISTONE = {"P", "D", "C", "A"}
_ROLE_ALLOWED_VOTI = {"P", "D", "C", "A"}  # "ALL" (allenatore) is filtered out


# =========================================================================
# Small helpers
# =========================================================================

def _canonicalize_team(raw: Optional[str]) -> Optional[str]:
    """Return the canonical team name or ``None`` if unknown."""
    if not raw:
        return None
    key = re.sub(r"[^a-zA-Z\s]", "", str(raw)).strip().lower()
    if key in _SERIE_A_LOWER:
        name = _SERIE_A_LOWER[key]
        return TEAM_ALIASES.get(name, name)
    return None


def _safe_int(v) -> int:
    if v is None:
        return 0
    if isinstance(v, bool):
        return int(v)
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(v))
        except (TypeError, ValueError):
            return 0


def _parse_voto(v) -> Tuple[Optional[float], bool]:
    """Parse a "Voto" cell. Returns ``(value_or_None, is_senza_voto)``.

    The Excel voto column may be:
      * a numeric type (e.g. ``5.5``)                → (5.5, False)
      * a string with a trailing ``*``  (e.g. "6*")  → (6.0, True)
      * empty / None                                 → (None, True)
    """
    if v is None:
        return None, True
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v), False
    s = str(v).strip()
    if not s:
        return None, True
    sv = "*" in s
    cleaned = s.replace("*", "").replace(",", ".").strip()
    try:
        return float(cleaned), sv
    except ValueError:
        return None, True


def _split_name(name_field: str) -> Tuple[str, str]:
    """Split a fantacalcio.it player label into ``(first_name, last_name)``.

    Formats encountered in practice:
      * ``Svilar``       → ("", "Svilar")
      * ``Martinez L.``  → ("L.", "Martinez")
      * ``Paz N.``       → ("N.", "Paz")
      * ``Martinez Jo.`` → ("Jo.", "Martinez")
    """
    parts = name_field.strip().split()
    if len(parts) >= 2 and re.fullmatch(r"[A-Z][a-z]?\.", parts[-1]):
        return parts[-1], " ".join(parts[:-1])
    return "", name_field.strip()


def _is_disclaimer_row(cells: tuple) -> bool:
    """A disclaimer/title row has all values None except column A which is a
    string containing one of the well-known markers."""
    if not cells:
        return True
    non_null = [c for c in cells if c is not None]
    if len(non_null) != 1:
        return False
    first = cells[0]
    if not isinstance(first, str):
        return False
    return any(m.lower() in first.lower() for m in _DISCLAIMER_MARKERS)


# =========================================================================
# LISTONE parser
# =========================================================================

def parse_listone_xlsx(xlsx_bytes: bytes) -> List[dict]:
    """Parse a "Quotazioni Fantacalcio" xlsx into a list of player dicts.

    Reads the ``Tutti`` sheet (falls back to the first sheet). Rows in the
    ``Ceduti`` (transferred out) sheet are intentionally ignored because those
    players are no longer in Serie A.

    The output shape matches ``scoreandlive._parse_listone_pdf`` so the
    ``sal_players`` insert logic stays untouched.
    """
    try:
        import openpyxl  # noqa: WPS433
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(f"openpyxl non installato: {e}") from e

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheet_name = "Tutti" if "Tutti" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    players: List[dict] = []
    seen_ids: set = set()

    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        if _is_disclaimer_row(row):
            continue

        # Header row: Id | R | RM | Nome | Squadra | ...
        c0 = row[0]
        if isinstance(c0, str) and c0.strip().lower() == "id":
            continue

        # Real rows have an integer id in column A
        try:
            fid = int(c0)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            continue

        role = str(row[1] or "").strip().upper()
        if role not in _ROLE_ALLOWED_LISTONE:
            continue

        name_field = str(row[3] or "").strip()
        team_raw = str(row[4] or "").strip()
        if not name_field or not team_raw:
            continue

        team = _canonicalize_team(team_raw) or team_raw
        if fid in seen_ids:
            continue
        seen_ids.add(fid)

        first_name, last_name = _split_name(name_field)
        players.append({
            "fanta_id": fid,
            "first_name": first_name,
            "last_name": last_name,
            "team": team,
            "role": role,
            "role_mantra": (str(row[2]).strip() if row[2] else None),
            "price_current": _safe_int(row[5]) if len(row) > 5 else 0,
            "price_initial": _safe_int(row[6]) if len(row) > 6 else 0,
        })

    return players


# =========================================================================
# VOTI parser
# =========================================================================

def parse_voti_xlsx(
    xlsx_bytes: bytes,
    sheet: str = "Fantacalcio",
) -> Tuple[Optional[int], List[dict], Dict[str, Any]]:
    """Parse a "Voti Fantacalcio Giornata X" xlsx.

    Returns a triple ``(matchday, rows, diagnostics)`` mirroring the
    ``_parse_voti_pdf`` return shape so the caller can keep the same
    dry-run/commit flow.

    * ``matchday`` is detected from the sheet title row (e.g. "38ª giornata")
    * Rows with ``Ruolo == "ALL"`` (allenatore) are filtered out
    * Team header rows (single string in column A) drive the current team
      context for the following player rows
    """
    try:
        import openpyxl  # noqa: WPS433
    except ImportError as e:  # pragma: no cover
        raise RuntimeError(f"openpyxl non installato: {e}") from e

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        # Fall back to the first sheet — the file may have been renamed
        sheet_name = wb.sheetnames[0]
    else:
        sheet_name = sheet
    ws = wb[sheet_name]

    matchday: Optional[int] = None
    current_team: Optional[str] = None
    rows: List[dict] = []
    seen_keys: set = set()

    # Diagnostics
    teams_seen: List[str] = []
    lines_scanned = 0
    excluded_all = 0
    excluded_no_team = 0
    excluded_unknown_role = 0
    sample_unmatched: List[str] = []

    for raw in ws.iter_rows(values_only=True):
        if not raw:
            continue
        # openpyxl in read_only mode may return rows padded with Nones; strip
        # trailing Nones for cleanliness but keep positional indexes.
        cells = list(raw)
        non_null = [c for c in cells if c is not None]
        if not non_null:
            continue
        lines_scanned += 1
        c0 = cells[0]

        # Try to detect matchday from any title row (first columns text)
        if matchday is None and isinstance(c0, str):
            m = _MATCHDAY_RE.search(c0)
            if m:
                try:
                    matchday = int(m.group(1))
                except ValueError:
                    pass

        # Team header row → single string in column A
        if isinstance(c0, str) and len(non_null) == 1:
            if _is_disclaimer_row(tuple(cells)):
                continue
            canon = _canonicalize_team(c0)
            if canon:
                current_team = canon
                if canon not in teams_seen:
                    teams_seen.append(canon)
            elif len(sample_unmatched) < 5:
                sample_unmatched.append(f"[team?] {str(c0)[:80]}")
            continue

        # Header row: "Cod." / "Ruolo" / "Nome" ...
        if isinstance(c0, str) and c0.strip().lower().startswith("cod"):
            continue

        # Player row: column A must be a numeric code
        try:
            code = int(c0)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            if isinstance(c0, str) and len(sample_unmatched) < 5:
                sample_unmatched.append(str(c0)[:100])
            continue

        role_raw = cells[1] if len(cells) > 1 else None
        role = str(role_raw or "").strip().upper()
        if role == "ALL":
            excluded_all += 1
            continue
        if role not in _ROLE_ALLOWED_VOTI:
            excluded_unknown_role += 1
            continue
        if not current_team:
            excluded_no_team += 1
            continue

        name = str(cells[2] or "").strip() if len(cells) > 2 else ""
        if not name:
            continue

        voto_val, sv_flag = _parse_voto(cells[3] if len(cells) > 3 else None)

        gf = _safe_int(cells[4] if len(cells) > 4 else 0)
        gs = _safe_int(cells[5] if len(cells) > 5 else 0)
        rp = _safe_int(cells[6] if len(cells) > 6 else 0)
        rs = _safe_int(cells[7] if len(cells) > 7 else 0)
        rf = _safe_int(cells[8] if len(cells) > 8 else 0)
        au = _safe_int(cells[9] if len(cells) > 9 else 0)
        amm = _safe_int(cells[10] if len(cells) > 10 else 0)
        esp = _safe_int(cells[11] if len(cells) > 11 else 0)
        ass = _safe_int(cells[12] if len(cells) > 12 else 0)

        key = (current_team, code)
        if key in seen_keys:
            continue
        seen_keys.add(key)

        rows.append({
            "matchday": matchday or 0,
            "team": current_team,
            "player_code": code,
            "player_name": name,
            "role": role,
            "voto": voto_val,
            "sv": sv_flag,
            "gf": gf,
            "gs": gs,
            "rp": rp,
            "rs": rs,
            "rf": rf,
            "au": au,
            "amm": amm,
            "esp": esp,
            "ass": ass,
            "total_goals": gf + rf,
        })

    diagnostics = {
        "sheet_used": sheet_name,
        "lines_scanned": lines_scanned,
        "teams_seen": teams_seen,
        "teams_seen_count": len(teams_seen),
        "excluded_all": excluded_all,
        "excluded_no_team": excluded_no_team,
        "excluded_unknown_role": excluded_unknown_role,
        "sample_unmatched": sample_unmatched,
        "matchday_detected": matchday,
    }
    return matchday, rows, diagnostics


# =========================================================================
# CLI smoke test (usage: python -m excel_parser voti.xlsx)
# =========================================================================

if __name__ == "__main__":  # pragma: no cover
    import sys, json

    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <path-to-xlsx> [listone|voti]")
        sys.exit(1)

    path = sys.argv[1]
    mode = sys.argv[2] if len(sys.argv) > 2 else "auto"
    with open(path, "rb") as fh:
        data = fh.read()

    if mode == "listone" or (mode == "auto" and "Quotazioni" in path):
        players = parse_listone_xlsx(data)
        print(f"Listone: {len(players)} players")
        by_team: Dict[str, int] = {}
        by_role: Dict[str, int] = {}
        for p in players:
            by_team[p["team"]] = by_team.get(p["team"], 0) + 1
            by_role[p["role"]] = by_role.get(p["role"], 0) + 1
        print("By team:", json.dumps(dict(sorted(by_team.items())), indent=2))
        print("By role:", json.dumps(dict(sorted(by_role.items())), indent=2))
        print("Sample:", json.dumps(players[:5], indent=2, ensure_ascii=False))
    else:
        md, rows, diag = parse_voti_xlsx(data)
        print(f"Voti giornata: {md} · {len(rows)} righe")
        print("Diagnostics:", json.dumps(diag, indent=2, ensure_ascii=False))
        print("Sample:", json.dumps(rows[:3], indent=2, ensure_ascii=False))
